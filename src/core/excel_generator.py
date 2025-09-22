"""
Generador de archivo Excel con columnas detalladas de fichadas - VERSIÓN FINAL
Incluye sedes, incidencias, headers corregidos y máximo 4 pares entrada/salida
RESTA 3 HORAS para convertir UTC a horario de Argentina
Colores informativos: Solo "Sí" destacados, resto en blanco
NUEVA FUNCIONALIDAD: Columna de Permisos Pedidos Aprobados
"""

import os
from datetime import datetime, timedelta
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from config.default_config import DEFAULT_CONFIG


class ExcelReportGeneratorDetailed:
    """Genera un Excel con columnas detalladas para cada entrada/salida - VERSIÓN FINAL."""
    
    def __init__(self):
        self.output_dir = os.path.expanduser(DEFAULT_CONFIG['output_directory'])
        self.filename_format = DEFAULT_CONFIG['filename_format']
        
        # Estilos predefinidos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        self.light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    def _format_hours_exact(self, hours_float: float) -> str:
        """Convierte horas decimales a string EXACTO sin redondeo"""
        if not hours_float or hours_float <= 0:
            return "0"

        total_minutes = round(hours_float * 60)
        hours = total_minutes // 60
        minutes = total_minutes % 60

        if minutes == 0:
            return f"{hours}"
        return f"{hours}:{minutes:02d}"

    def _format_turno_from_timeslots(self, time_slots: List[Dict]) -> str:
        """Formatea los timeSlots como 'HH:MM-HH:MM'"""
        if not time_slots or not time_slots[0]:
            return ""
        
        slot = time_slots[0]
        start_time = slot.get('startTime', '')
        end_time = slot.get('endTime', '')
        
        if start_time and end_time:
            return f"{start_time}-{end_time}"
        return ""

    def _process_entries_for_columns(self, entries: List[Dict]) -> Dict:
        """
        Procesa las entries y las organiza en PARES de entrada/salida
        Máximo 4 pares (8 entradas total)
        RESTA 3 HORAS para convertir UTC a horario de Argentina
        """
        # Ordenar entries por tiempo
        sorted_entries = sorted(entries, key=lambda x: x.get('time', ''))
        
        # Agrupar en pares entrada/salida
        pairs = []
        current_start = None
        
        for entry in sorted_entries:
            entry_type = entry.get('type')
            time_str = entry.get('time', '')
            comment = entry.get('comment', '')
            site_name = entry.get('site', {}).get('name', '') if entry.get('site') else ''
            
            # Extraer hora en formato HH:MM y RESTAR 3 HORAS
            try:
                if time_str:
                    dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
                    # RESTAR 3 HORAS para convertir UTC a Argentina
                    dt_argentina = dt - timedelta(hours=3)
                    time_formatted = dt_argentina.strftime('%H:%M')
                else:
                    time_formatted = ''
            except:
                time_formatted = ''
            
            if entry_type == 'START':
                current_start = {
                    'time': time_formatted,
                    'comment': comment or '',
                    'site': site_name,
                    'raw_entry': entry
                }
            elif entry_type == 'END' and current_start:
                # Completar el par
                pairs.append({
                    'entrada_hora': current_start['time'],
                    'entrada_comment': current_start['comment'],
                    'entrada_site': current_start['site'],
                    'salida_hora': time_formatted,
                    'salida_comment': comment or '',
                    'salida_site': site_name
                })
                current_start = None
        
        # Si queda un START sin END
        if current_start:
            pairs.append({
                'entrada_hora': current_start['time'],
                'entrada_comment': current_start['comment'],
                'entrada_site': current_start['site'],
                'salida_hora': '',
                'salida_comment': '',
                'salida_site': ''
            })
        
        # Limitar a máximo 4 pares
        pairs = pairs[:4]
        
        return {
            'pairs': pairs,
            'total_pairs': len(pairs)
        }

    def _determine_max_pairs(self, processed_data: Dict) -> int:
        """Determina el número máximo de pares entrada/salida (máximo 4)"""
        max_pairs = 0
        
        for employee_data in processed_data.values():
            for daily_record in employee_data['daily_data']:
                entries = daily_record.get('raw_entries', [])
                if entries:
                    entry_data = self._process_entries_for_columns(entries)
                    max_pairs = max(max_pairs, entry_data['total_pairs'])
        
        return min(max(max_pairs, 1), 4)  # Mínimo 1, máximo 4
    
    def generate_report(self, processed_data: Dict, start_date: str, end_date: str, 
                       output_filename: str = None) -> str:
        """
        Genera el reporte Excel con columnas detalladas - VERSIÓN FINAL
        """
        try:
            # Determinar número máximo de pares (máximo 4)
            max_pairs = self._determine_max_pairs(processed_data)
            
            # Crear workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Crear hoja con columnas detalladas
            self._create_detailed_sheet(wb, processed_data, start_date, end_date, max_pairs)
            
            # Generar nombre de archivo
            if not output_filename:
                output_filename = self.filename_format.format(
                    start_date=start_date.replace('-', ''),
                    end_date=end_date.replace('-', '')
                ).replace('.xlsx', '_detallado.xlsx')
            
            # Asegurar que el directorio existe
            os.makedirs(self.output_dir, exist_ok=True)
            
            # Guardar archivo
            filepath = os.path.join(self.output_dir, output_filename)
            wb.save(filepath)
            
            print(f"Reporte Excel detallado generado: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error generando reporte Excel detallado: {str(e)}")
            raise e
    
    def _create_detailed_sheet(self, wb: Workbook, processed_data: Dict, 
                             start_date: str, end_date: str, max_pairs: int):
        """Crea la hoja con columnas detalladas de fichadas - VERSIÓN FINAL CON PERMISOS"""
        ws = wb.create_sheet("Fichadas Detalladas")
        
        # Título
        ws['A1'] = f"DETALLE DE FICHADAS - COLUMNAS EXPANDIDAS CON PERMISOS"
        ws['A2'] = f"Período: {start_date} al {end_date}"
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Aplicar estilo al título
        for row in range(1, 4):
            ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Construir headers dinámicamente - VERSIÓN FINAL CON PERMISOS
        base_headers = [
            'ID Empleado', 'Nombre', 'Apellido', 'Fecha', 'Día',
            'Turno Programado', 'Permisos Pedidos Aprobados', 'Tardanza', 'Trabajo Menos'
        ]
        
        # Headers dinámicos para pares entrada/salida
        entry_headers = []
        for i in range(1, max_pairs + 1):
            entry_headers.extend([
                f'Sede Ingreso {i}',
                f'Hora de Ingreso {i}',
                f'Comentario Ingreso {i}',
                f'Sede Salida {i}',
                f'Hora de Salida {i}',
                f'Comentario Salida {i}'
            ])
        
        # Headers de resumen
        summary_headers = [
            'Horas Trabajadas', 'Horas Regulares', 'Horas Extra 50%', 
            'Horas Extra 100%', 'Horas Nocturnas', 'Horas Pendientes',
            'Es Feriado', 'Nombre Feriado', 'Tiene Licencia', 
            'Tipo Licencia', 'Observaciones'
        ]
        
        # Combinar todos los headers
        all_headers = base_headers + entry_headers + summary_headers
        
        # Escribir headers
        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment
        
        # Calcular índices de columnas importantes
        COL_TURNO = 6
        COL_PERMISOS = 7  # NUEVA COLUMNA
        COL_TARDANZA = 8
        COL_TRABAJO_MENOS = 9
        COL_ENTRIES_START = 10  # Ajustado por la nueva columna
        COL_SUMMARY_START = COL_ENTRIES_START + (max_pairs * 6)
        
        # Escribir datos
        row = 6
        for employee_data in processed_data.values():
            employee_info = employee_data['employee_info']
            
            for daily_record in employee_data['daily_data']:
                # Obtener incidencias
                incidences = daily_record.get('incidences', [])
                
                # Procesar incidencias
                has_late = 'LATE' in incidences
                has_underworked = 'UNDERWORKED' in incidences
                
                # Obtener permiso para esta fila
                permission_text = daily_record.get('permission_request', '')
                
                # Datos base CON PERMISOS
                base_data = [
                    employee_info.get('employeeInternalId', ''),
                    employee_info.get('firstName', ''),
                    employee_info.get('lastName', ''),
                    daily_record['date'],
                    daily_record['day_of_week'],
                    self._format_turno_from_timeslots(daily_record.get('time_slots', [])),
                    permission_text,  # NUEVA COLUMNA DE PERMISOS
                    'Sí' if has_late else 'No',
                    'Sí' if has_underworked else 'No'
                ]
                
                # Procesar entries para pares
                raw_entries = daily_record.get('raw_entries', [])
                entry_data = self._process_entries_for_columns(raw_entries)
                
                # Datos de pares entrada/salida
                entry_columns_data = []
                for i in range(max_pairs):
                    if i < len(entry_data['pairs']):
                        pair = entry_data['pairs'][i]
                        entry_columns_data.extend([
                            pair['entrada_site'],
                            pair['entrada_hora'],
                            pair['entrada_comment'],
                            pair['salida_site'],
                            pair['salida_hora'],
                            pair['salida_comment']
                        ])
                    else:
                        entry_columns_data.extend(['', '', '', '', '', ''])
                
                # Datos de resumen
                observations = []
                if daily_record['is_holiday']:
                    observations.append(f"Feriado: {daily_record['holiday_name'] or 'N/A'}")
                if daily_record['has_time_off']:
                    observations.append(f"Licencia: {daily_record['time_off_name'] or 'N/A'}")
                if daily_record['pending_hours'] > 0:
                    observations.append(f"{self._format_hours_exact(daily_record['pending_hours'])} pendientes")
                if has_late:
                    observations.append("Tardanza")
                if has_underworked:
                    observations.append("Trabajo insuficiente")
                if permission_text:
                    observations.append("Permiso solicitado")
                
                summary_data = [
                    self._format_hours_exact(daily_record['hours_worked']),
                    self._format_hours_exact(daily_record['regular_hours']),
                    self._format_hours_exact(daily_record['extra_hours_50']),
                    self._format_hours_exact(daily_record['extra_hours_100']),
                    self._format_hours_exact(daily_record['night_hours']),
                    self._format_hours_exact(daily_record['pending_hours']),
                    'Sí' if daily_record['is_holiday'] else 'No',
                    daily_record['holiday_name'] or '',
                    'Sí' if daily_record['has_time_off'] else 'No',
                    daily_record['time_off_name'] or '',
                    ', '.join(observations) if observations else ''
                ]
                
                # Combinar todos los datos
                complete_row = base_data + entry_columns_data + summary_data
                
                # Escribir la fila
                for col, value in enumerate(complete_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.thin_border
                    
                    # Aplicar colores según el tipo de columna
                    if col == COL_TURNO:
                        cell.fill = self.white_fill  # Blanco
                    elif col == COL_PERMISOS:
                        # Color celeste para la columna de permisos
                        if value and value.strip():
                            cell.fill = self.light_blue_fill  # Celeste si hay permiso
                        else:
                            cell.fill = self.white_fill  # Blanco si no hay permiso
                    elif col == COL_TARDANZA:
                        if has_late and value == 'Sí':
                            cell.fill = self.red_fill  # Rojo para Sí
                        else:
                            cell.fill = self.green_fill  # Verde para No
                    elif col == COL_TRABAJO_MENOS:
                        cell.fill = self.green_fill  # Verde siempre (Sí y No)
                    elif COL_ENTRIES_START <= col < COL_SUMMARY_START:
                        # Todas las columnas de entries en blanco
                        cell.fill = self.white_fill
                    elif col >= COL_SUMMARY_START:
                        # Columnas de resumen
                        summary_col_offset = col - COL_SUMMARY_START
                        if summary_col_offset == 6:  # Es Feriado
                            if value == 'Sí':
                                cell.fill = self.green_fill  # Verde para Sí
                            else:
                                cell.fill = self.white_fill  # Blanco para No
                        elif summary_col_offset == 8:  # Tiene Licencia
                            if value == 'Sí':
                                cell.fill = self.green_fill  # Verde para Sí
                            else:
                                cell.fill = self.white_fill  # Blanco para No
                        else:
                            # Todas las demás columnas de resumen en blanco
                            cell.fill = self.white_fill
                
                row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(all_headers) + 1):
            if col <= 6:  # Columnas base hasta turno programado
                ws.column_dimensions[get_column_letter(col)].width = 14
            elif col == COL_PERMISOS:  # Columna de permisos más ancha
                ws.column_dimensions[get_column_letter(col)].width = 25
            elif col in [COL_TARDANZA, COL_TRABAJO_MENOS]:  # Tardanza y trabajo menos
                ws.column_dimensions[get_column_letter(col)].width = 14
            elif COL_ENTRIES_START <= col < COL_SUMMARY_START:
                # Columnas de entries
                relative_col = (col - COL_ENTRIES_START) % 6
                if relative_col in [0, 3]:  # Sedes
                    ws.column_dimensions[get_column_letter(col)].width = 15
                elif relative_col in [1, 4]:  # Horas
                    ws.column_dimensions[get_column_letter(col)].width = 12
                else:  # Comentarios
                    ws.column_dimensions[get_column_letter(col)].width = 18
            else:  # Columnas de resumen
                ws.column_dimensions[get_column_letter(col)].width = 12


# Alias para compatibilidad con código existente
ExcelReportGenerator = ExcelReportGeneratorDetailed